"""Read TUM's published VerSe demographics spreadsheet.

The spreadsheet (``configs/verse_demographics.xlsx``) is the authoritative
source for patient identity in VerSe.  Schema:

    column A: subject               canonical patient ID (e.g. "verse014", "verse400", "gl003")
    column B: split                 either empty (single-series patient) or
                                    the original MICCAI series ID for this row
                                    (e.g. "verse090", "verse155" for the two
                                    scans of patient "verse400")
    column C: CT_image_series       position-in-series indicator ("1 of 1",
                                    "1 of 2", "2 of 3", etc.)
    column D: verse_2019            1 if subject appears in VerSe 2019 release
    column E: verse_2020            1 if subject appears in VerSe 2020 release
    column F: sex (0= f, 1= m)
    column G: age

Total rows: 374 (one per image series across both releases).
Unique patients: 355.

This module exposes a normalised representation: one ``DemographicRow`` per
image series, with ``series_id`` always set to the bare MICCAI filename stem
(``verse014``, ``verse090``, ``gl003``) and ``patient_id`` to the canonical
group ID.  Sibling scans of the same patient share ``patient_id``.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path

import openpyxl

log = logging.getLogger("verse.demographics")


@dataclass(frozen=True)
class DemographicRow:
    """One row of TUM's demographic table, normalised."""
    series_id:        str          # MICCAI filename stem (e.g. "verse090", "gl003")
    patient_id:       str          # canonical patient group ID (e.g. "verse400")
    position:         str          # "1 of 1", "1 of 2", "2 of 3", etc.
    in_v19:           bool
    in_v20:           bool
    sex:              str          # "F" | "M" | "?"
    age:              int | None


def _normalize(value) -> str:
    """Trim and stringify a cell, treating None as ''."""
    return str(value).strip() if value is not None else ""


def load_demographics(xlsx_path: Path) -> list[DemographicRow]:
    """Parse the demographics spreadsheet; return one ``DemographicRow`` per image series.

    The spreadsheet's column B is overloaded:
      - For single-series patients, it's empty.
      - For multi-series patients, it holds the original MICCAI series ID of
        the specific scan represented by this row.

    Both cases produce a row with ``series_id`` set to the unique series
    identifier and ``patient_id`` set to the canonical grouping key.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    raw = list(ws.iter_rows(values_only=True))

    if not raw or len(raw) < 2:
        raise ValueError(f"{xlsx_path}: spreadsheet appears empty")

    header = raw[0]
    expected = ("subject", "split", "CT_image_series",
                "verse_2019", "verse_2020")
    if tuple(str(h).strip() for h in header[:5]) != expected:
        raise ValueError(
            f"{xlsx_path}: unexpected header {header[:5]!r}; expected {expected!r}"
        )

    rows: list[DemographicRow] = []
    skipped = 0
    for raw_row in raw[1:]:
        canon = _normalize(raw_row[0])
        col_b = _normalize(raw_row[1])
        position = _normalize(raw_row[2])
        in_v19 = raw_row[3] == 1
        in_v20 = raw_row[4] == 1

        # The summary row at the bottom of the spreadsheet has a numeric
        # subject (e.g. 374); skip it.
        if not canon or not canon.startswith(("verse", "gl")):
            skipped += 1
            continue

        # Determine the unique series ID for this row.
        # If col_b starts with "verse" or "gl", it's the original MICCAI series ID
        # of this particular scan of a multi-series patient.  Otherwise, the
        # row's series ID is the same as its canonical patient ID.
        if col_b.startswith(("verse", "gl")):
            series_id = col_b
            patient_id = canon
        else:
            series_id = canon
            patient_id = canon

        sex_val = raw_row[5]
        if sex_val == 0:
            sex = "F"
        elif sex_val == 1:
            sex = "M"
        else:
            sex = "?"

        age_val = raw_row[6]
        age = int(age_val) if isinstance(age_val, (int, float)) else None

        rows.append(DemographicRow(
            series_id=series_id, patient_id=patient_id, position=position,
            in_v19=in_v19, in_v20=in_v20, sex=sex, age=age,
        ))

    if skipped:
        log.debug("Skipped %d non-subject row(s) from %s", skipped, xlsx_path)

    log.info("Loaded %d demographic rows from %s", len(rows), xlsx_path)
    return rows


def index_by_series(rows: list[DemographicRow]) -> dict[str, DemographicRow]:
    """Return a mapping from series_id to its single ``DemographicRow``.

    Raises ``ValueError`` if duplicate series IDs are encountered, which
    would indicate either a corrupted spreadsheet or a bug here.
    """
    out: dict[str, DemographicRow] = {}
    for r in rows:
        if r.series_id in out:
            raise ValueError(
                f"Duplicate series_id {r.series_id!r} in demographics; "
                "spreadsheet should have one row per image series."
            )
        out[r.series_id] = r
    return out


def patients(rows: list[DemographicRow]) -> dict[str, list[DemographicRow]]:
    """Group rows by ``patient_id``; useful for multi-series accounting."""
    out: dict[str, list[DemographicRow]] = {}
    for r in rows:
        out.setdefault(r.patient_id, []).append(r)
    return out
